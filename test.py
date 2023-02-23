import re
import openpyxl



def search_txt():
    file = open("text-reader.txt", "r")
    text = file.read()
    word1 = re.compile(r"Business\s+(\w+\s+\w+)")
    word2 = re.compile(r"Termen plata:\s+(\d{2}[/-]\d{2}[/-]\d{4})")
    word3 = re.compile(r"buc. 1\s+(\w+)")

    match1 = word1.search(text)
    if match1:
        next2w1 = match1.group(1)
    else:
        next2w1 = None
    print(next2w1, word1)
    # print("Next two words after word1: ", next2w1)

    match2 = word2.search(text)
    if match2:
        next2w2 = match2.group(1)
    else:
        next2w2 = None
    print(next2w2, word2)
    # print("Next two words after word1: ", next2w1)

    match3 = word3.search(text)
    if match3:
        next2w3 = match3.group(1)
    else:
        next2w3 = None
    print(next2w3, word3)
    # print("Next two words after word1: ", next2w1)

    wb = openpyxl.load_workbook("ff.xlsx")
    ws = wb.active
    last_row = ws.max_row + 1
    words = (next2w1, next2w2, next2w3)
    for i in range(len(words)):
        col1 = i+1
        col2 = i+2
        #col3 = i+3

        ws.cell(row=last_row, column=1, value=words[0])
        ws.cell(row=last_row, column=4, value=words[1] if col1 < len(words) else None)
        ws.cell(row=last_row, column=3, value=words[2] if col2 < len(words) else None)
        ws.cell(row=last_row, column=2, value="=today()+365")
        last_row += 1
    wb.save("ff.xlsx")

search_txt()